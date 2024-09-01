import customtkinter as ctk
from tkinter import messagebox
import openpyxl
import os

class EmployeesPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.employee_list = []

        # Load existing employees from the Excel sheet
        self.load_employees_from_excel()

        # Add Employee Button
        self.add_employee_button = ctk.CTkButton(self, text="Add Employee", command=self.add_employee)
        self.add_employee_button.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        # Title: Employee Data
        self.title_label = ctk.CTkLabel(self, text="Employee Data", font=("Arial", 26))  # Larger title font
        self.title_label.grid(row=0, column=1, padx=10, pady=10, sticky="n")
        self.grid_columnconfigure(1, weight=1)  # Center the title

        # Scrollable Frame for Employee List
        self.employee_frame = ctk.CTkScrollableFrame(self, width=800, height=400)
        self.employee_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.employee_frame.grid_columnconfigure(0, weight=1)  # Expand the frame to the edge
        self.grid_rowconfigure(1, weight=1)

        # Display employees
        self.display_employees()

    def calculate_column_widths(self):
        # Set fixed column widths for each data point
        column_widths = [
            50,   # Number (Employee #)
            200,  # Arabic Name
            250,  # English Name
            120,  # ID Number
            120,  # Phone Number
            170,  # Email
            100,  # Entry Date
            100,  # Exit Date
            80,   # Basic Salary
            50,   # Housing Allowance
            50    # Transportation Allowance
        ]

        return column_widths[::-1]  # Reverse to match the right-to-left order

    def display_column_headers(self, column_widths):
        headers = ["#", "الإســــــــــم", "Name", "رقم الهوية", "رقم الجوال", 
                   "البريد الالكتروني", "تاريخ الدخول", "تاريخ الخروج", 
                   "الراتب أساسي", "بدل سكن", "بدل مواصلات"]

        # Display headers from right to left with fixed widths
        for i, header in enumerate(headers[::-1]):
            header_label = ctk.CTkLabel(self.employee_frame, text=header, font=("Arial", 16), anchor="center", width=column_widths[i])
            header_label.grid(row=0, column=i+1, padx=10, pady=5, sticky="e")

        # Add a blank label to align the headers properly
        blank_label = ctk.CTkLabel(self.employee_frame, text="", width=column_widths[0])
        blank_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")

    def display_employees(self):
        for widget in self.employee_frame.winfo_children():
            widget.destroy()

        # Calculate the necessary column widths (now fixed)
        column_widths = self.calculate_column_widths()

        # Display the column headers
        self.display_column_headers(column_widths)

        for i, employee in enumerate(self.employee_list, start=1):
            # Edit button on the far left
            edit_button = ctk.CTkButton(self.employee_frame, text="Edit", font=("Arial", 14),
                                        command=lambda e=employee: self.show_employee_popup(e))
            edit_button.grid(row=i, column=0, padx=10, pady=5, sticky="w")

            # Display each employee's information as a row from right to left
            entry_widths = {
                "#": 50,
                "الإســــــــــم": 200,
                "Name": 250,
                "رقم الهوية": 120,
                "رقم الجوال": 120,
                "البريد الالكتروني": 170,
                "تاريخ الدخول": 100,
                "تاريخ الخروج": 100,
                "الراتب أساسي": 80,
                "بدل سكن": 50,
                "بدل مواصلات": 50,
            }

            for j, (key, value) in enumerate(employee.items()):
                employee_info_label = ctk.CTkEntry(self.employee_frame, font=("Arial", 14), width=entry_widths[key], justify="center")
                employee_info_label.insert(0, value)
                employee_info_label.configure(state="readonly")  # Make the text read-only but selectable
                employee_info_label.grid(row=i, column=len(employee) - j, padx=5, pady=5, sticky="e")


    def add_employee(self):
        # Show the Add/Edit Employee Popup
        self.show_employee_popup()

    def show_employee_popup(self, employee_data=None):
        # Create the pop-up window
        popup = ctk.CTkToplevel(self)
        popup.title("Add/Edit Employee")
        popup.geometry("400x600")  # Set a taller size for the pop-up
        popup.resizable(False, False)  # Lock the pop-up size
        popup.transient(self)  # Ensure pop-up stays on top
        popup.grab_set()  # Make the pop-up modal (prevents interaction with main window)

        # Employee Data Fields
        labels = ["#", "الإســــــــــم", "Name", "رقم الهوية", "رقم الجوال", 
                "البريد الالكتروني", "تاريخ الدخول", "تاريخ الخروج", 
                "الراتب أساسي", "بدل سكن", "بدل مواصلات"]
        entries = {}

        # Determine the next employee number if creating a new employee
        if employee_data is None:
            employee_number = str(len(self.employee_list) + 1)
        else:
            employee_number = employee_data.get("#", "")

        # Correct width mapping for each text box
        entry_widths = {
            "#": 50,
            "الإســــــــــم": 200,
            "Name": 250,
            "رقم الهوية": 120,
            "رقم الجوال": 120,
            "البريد الالكتروني": 170,
            "تاريخ الدخول": 100,
            "تاريخ الخروج": 100,
            "الراتب أساسي": 80,
            "بدل سكن": 50,
            "بدل مواصلات": 50,
        }

        # Increase font size for better readability and adjust text boxes
        for i, label_text in enumerate(labels):
            label = ctk.CTkLabel(popup, text=label_text, font=("Arial", 14))  # Increased font size
            label.grid(row=i, column=0, padx=10, pady=10, sticky="e")

            entry = ctk.CTkEntry(popup, font=("Arial", 14), width=entry_widths[label_text])  # Set width based on label
            entry.grid(row=i, column=1, padx=10, pady=10, sticky="w", ipadx=10)  # Fill width

            if label_text == "#":
                entry.insert(0, employee_number)
                entry.configure(state="readonly")
            elif employee_data:
                entry.insert(0, employee_data.get(label_text, ""))

            entries[label_text] = entry

        # Save Button
        save_button = ctk.CTkButton(popup, text="Save", command=lambda: self.save_employee_data(popup, entries))
        save_button.grid(row=len(labels), column=0, columnspan=2, pady=20)

    def save_employee_data(self, popup, entries):
        # Extract data from entries
        employee_data = {label: entry.get() for label, entry in entries.items()}
        employee_number = employee_data["#"]

        # Check if editing an existing employee or adding a new one
        is_new_employee = True
        for index, existing_employee in enumerate(self.employee_list):
            if existing_employee["#"] == employee_number:
                # Update the existing employee in the list
                self.employee_list[index] = employee_data
                is_new_employee = False
                break

        # Save to Excel
        self.save_to_excel(employee_data, is_new_employee)

        # Refresh the employee display
        self.display_employees()

        # Close the popup
        popup.destroy()

    def save_to_excel(self, employee_data, is_new_employee):
        # Check if file exists, if not create it
        file_path = os.path.normpath(self.controller.excel_file_path)
        file_exists = os.path.exists(file_path)

        if not file_exists:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "EmployeeData"
            headers = ["#", "الإســــــــــم", "Name", "رقم الهوية", "رقم الجوال",
                       "البريد الالكتروني", "تاريخ الدخول", "تاريخ الخروج",
                       "الراتب أساسي", "بدل سكن", "بدل مواصلات"]
            sheet.append(headers)
        else:
            workbook = openpyxl.load_workbook(file_path)
            if "EmployeeData" not in workbook.sheetnames:
                sheet = workbook.create_sheet("EmployeeData")
                headers = ["#", "الإســــــــــم", "Name", "رقم الهوية", "رقم الجوال",
                           "البريد الالكتروني", "تاريخ الدخول", "تاريخ الخروج",
                           "الراتب أساسي", "بدل سكن", "بدل مواصلات"]
                sheet.append(headers)
            else:
                sheet = workbook["EmployeeData"]

        if is_new_employee:
            # Append employee data based on the correct order of headers
            data_row = [employee_data.get(header, "") for header in ["#", "الإســــــــــم", "Name", "رقم الهوية",
                                                                     "رقم الجوال", "البريد الالكتروني", "تاريخ الدخول",
                                                                     "تاريخ الخروج", "الراتب أساسي", "بدل سكن", "بدل مواصلات"]]
            sheet.append(data_row)
        else:
            # Update the existing employee data
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if str(row[0].value) == employee_data["#"]:
                    for idx, header in enumerate(["#", "الإســــــــــم", "Name", "رقم الهوية",
                                                  "رقم الجوال", "البريد الالكتروني", "تاريخ الدخول",
                                                  "تاريخ الخروج", "الراتب أساسي", "بدل سكن", "بدل مواصلات"]):
                        row[idx].value = employee_data.get(header, "")
                    break

        workbook.save(file_path)
        messagebox.showinfo("Success", "Employee data saved successfully!")

    def load_employees_from_excel(self):
        file_path = os.path.normpath(self.controller.excel_file_path)
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            if "EmployeeData" in workbook.sheetnames:
                sheet = workbook["EmployeeData"]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    employee_data = {
                        "#": row[0],
                        "الإســــــــــم": row[1],
                        "Name": row[2],
                        "رقم الهوية": row[3],
                        "رقم الجوال": row[4],
                        "البريد الالكتروني": row[5],
                        "تاريخ الدخول": row[6],
                        "تاريخ الخروج": row[7],
                        "الراتب أساسي": row[8],
                        "بدل سكن": row[9],
                        "بدل مواصلات": row[10],
                    }
                    self.employee_list.append(employee_data)
