import customtkinter as ctk
from tkinter import messagebox, Toplevel
import openpyxl
import os

class SalariesPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        # Initialize row_checkbox_dict to track checkbox states for each row
        self.row_checkbox_dict = {}

        # Year Dropdown
        self.year_var = ctk.StringVar(value="--Year--")
        self.year_dropdown = ctk.CTkComboBox(self, variable=self.year_var, values=[str(year) for year in range(2020, 2031)])
        self.year_dropdown.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        # Dash label
        self.dash_label = ctk.CTkLabel(self, text="-", font=("Arial", 16))
        self.dash_label.grid(row=0, column=1, padx=10, pady=10)

        # Month Dropdown
        self.month_var = ctk.StringVar(value="--Month--")
        self.month_dropdown = ctk.CTkComboBox(self, variable=self.month_var, values=[f"{month:02d}" for month in range(1, 13)] + ["bonus"])
        self.month_dropdown.grid(row=0, column=2, padx=10, pady=10, sticky="w")

        # Open Button
        self.open_button = ctk.CTkButton(self, text="Open", command=self.load_salary_sheet)
        self.open_button.grid(row=0, column=3, padx=10, pady=10, sticky="w")

        # Title: Monthly Salaries
        self.title_label = ctk.CTkLabel(self, text="Monthly Salaries", font=("Arial", 26))
        self.title_label.grid(row=0, column=4, padx=10, pady=10, sticky="n")
        self.grid_columnconfigure(4, weight=1)  # Center the title across the entire page

        # Preview Last Month Button
        self.preview_button = ctk.CTkButton(self, text="Preview Last Month", command=self.preview_last_month)
        self.preview_button.grid(row=0, column=5, padx=10, pady=10, sticky="e")

        # Save Changes Button
        self.save_button = ctk.CTkButton(self, text="Save Changes", command=self.save_changes)
        self.save_button.grid(row=0, column=6, padx=10, pady=10, sticky="e")

        # Scrollable Frame for Salary Data (empty area for now)
        self.salary_frame = ctk.CTkScrollableFrame(self, width=800, height=400)
        self.salary_frame.grid(row=1, column=0, columnspan=7, padx=10, pady=10, sticky="nsew")
        self.salary_frame.grid_columnconfigure(0, weight=1)  # Expand the frame to the edge
        self.grid_rowconfigure(1, weight=1)

    def preview_last_month(self):
        # Create the popup window
        self.popup = ctk.CTkToplevel(self)
        self.popup.title("Preview Last Month")
        self.popup.geometry("1000x600")

        # Apply the same appearance mode to the popup window
        current_mode = ctk.get_appearance_mode()
        ctk.set_appearance_mode(current_mode)

        # Ensure the popup stays on top
        self.popup.transient(self)  # Set the popup as a transient window for the main window
        self.popup.lift()  # Lift it above all windows
        self.popup.grab_set()  # Ensure all events are directed to the popup
        self.popup.focus_force()  # Force focus on the popup window

        # Create the same dropdowns for year and month
        year_var = ctk.StringVar(value="--Year--")
        year_dropdown = ctk.CTkComboBox(self.popup, variable=year_var, values=[str(year) for year in range(2020, 2031)])
        year_dropdown.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        dash_label = ctk.CTkLabel(self.popup, text="-", font=("Arial", 16))
        dash_label.grid(row=0, column=1, padx=10, pady=10)

        month_var = ctk.StringVar(value="--Month--")
        month_dropdown = ctk.CTkComboBox(self.popup, variable=month_var, values=[f"{month:02d}" for month in range(1, 13)] + ["bonus"])
        month_dropdown.grid(row=0, column=2, padx=10, pady=10, sticky="w")

        open_button = ctk.CTkButton(self.popup, text="Open", command=lambda: self.load_popup_sheet(year_var.get(), month_var.get()))
        open_button.grid(row=0, column=3, padx=10, pady=10, sticky="w")

        # Title: Monthly Salaries
        title_label = ctk.CTkLabel(self.popup, text="Preview Last Month", font=("Arial", 26))
        title_label.grid(row=0, column=4, padx=10, pady=10, sticky="n")
        self.popup.grid_columnconfigure(4, weight=1)

        # Scrollable Frame for the Preview Data
        self.preview_frame = ctk.CTkScrollableFrame(self.popup, width=800, height=400)
        self.preview_frame.grid(row=1, column=0, columnspan=7, padx=10, pady=10, sticky="nsew")
        self.preview_frame.grid_columnconfigure(0, weight=1)
        self.popup.grid_rowconfigure(1, weight=1)

        # Automatically load the previous month if available
        self.auto_load_previous_month(year_var, month_var)

    def load_popup_sheet(self, year, month):
        # Code to load the sheet in the popup (read-only mode)
        sheet_name = f"{year}-{month}"
        print(f"Loading sheet in popup: {sheet_name}")  # Debug print

        # Check if the Excel file exists, if not create it
        file_path = os.path.normpath(self.controller.excel_file_path)
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)

            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self.display_popup_sheet(sheet)
            else:
                messagebox.showwarning("Sheet Not Found", f"The sheet '{sheet_name}' does not exist.")
        else:
            messagebox.showerror("File Error", "The Excel file does not exist.")

    def display_popup_sheet(self, sheet):
        # Clear existing content in the popup
        for widget in self.preview_frame.winfo_children():
            widget.destroy()

        # Define column headers and widths
        headers = ["#", "الإســــــــــم", "النوع", "الراتب", "أخرى", "سداد السلف", "اضافي", 
                   "خصم", "ايام الغياب", "خصم الغياب", "السلف المتبقي", "الصافي", 
                   "تاريخ صدور الراتب", "ملاحظات"]
        column_widths = [50, 200, 50, 150, 100, 100, 100, 100, 50, 100, 150, 150, 100, 250]

        # Display the sheet content in read-only mode
        for j, (header, width) in enumerate(zip(headers[::-1], column_widths[::-1]), start=1):
            header_label = ctk.CTkLabel(self.preview_frame, text=header, font=("Arial", 16), justify="center", width=width)
            header_label.grid(row=1, column=j, padx=5, pady=5, sticky="n")

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            for j, (value, width) in enumerate(zip(reversed(row), column_widths[::-1]), start=1):
                value = "" if value is None else str(value)
                salary_info_label = ctk.CTkEntry(self.preview_frame, font=("Arial", 16), justify="center", width=width)
                salary_info_label.insert(0, value)
                salary_info_label.configure(state="readonly")
                salary_info_label.grid(row=i, column=j, padx=5, pady=5, sticky="e")

        # Adjust column configuration for better layout
        for j in range(len(headers) + 1):
            self.preview_frame.grid_columnconfigure(j, weight=1)
            self.preview_frame.grid_columnconfigure(len(headers) + 1, weight=0)  # Ensure no extra space

    def auto_load_previous_month(self, year_var, month_var):
        # Logic to automatically load the previous month if available
        current_year = int(self.year_var.get())
        current_month = self.month_var.get()

        if current_month == "01":
            previous_year = str(current_year - 1)
            previous_month = "12"
        elif current_month == "bonus":
            previous_year = str(current_year)
            previous_month = "12"
        else:
            previous_year = str(current_year)
            previous_month = f"{int(current_month) - 1:02d}"

        year_var.set(previous_year)
        month_var.set(previous_month)
        self.load_popup_sheet(previous_year, previous_month)
        
    def load_salary_sheet(self, event=None):
        print("Loading salary sheet...")  # Debug print
        # Get the selected year and month
        year = self.year_var.get()
        month = self.month_var.get()

        # Check if both selections are made
        if year != "--Year--" and month != "--Month--":
            if month == "bonus":
                sheet_name = f"{year}-bonus"
            else:
                sheet_name = f"{year}-{month}"
            
            self.selected_sheet_name = sheet_name  # Store the selected sheet name
            self.display_salary_sheet(sheet_name)
        else:
            messagebox.showwarning("Selection Error", "Please select both Year and Month.")

    def display_salary_sheet(self, sheet_name):
        print(f"Displaying sheet: {sheet_name}")  # Debug print
        # Clear existing content
        for widget in self.salary_frame.winfo_children():
            widget.destroy()

        # Check if the Excel file exists, if not create it
        file_path = os.path.normpath(self.controller.excel_file_path)
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            workbook.save(file_path)
        else:
            workbook = openpyxl.load_workbook(file_path)

        # Define column headers
        headers = ["#", "الإســــــــــم", "النوع", "الراتب", "أخرى", "سداد السلف", "اضافي", 
                "خصم", "ايام الغياب", "خصم الغياب", "السلف المتبقي", "الصافي", 
                "تاريخ صدور الراتب", "ملاحظات"]

        # Define column widths
        column_widths = [50, 200, 50, 150, 100, 100, 100, 100, 50, 100, 150, 150, 100, 250]

        # Display the column headers
        for j, (header, width) in enumerate(zip(headers[::-1], column_widths[::-1]), start=1):
            header_label = ctk.CTkLabel(self.salary_frame, text=header, font=("Arial", 16), justify="center", width=width)
            header_label.grid(row=1, column=j, padx=5, pady=5, sticky="n")

        # Check if the sheet exists, if not create it
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(headers)

            # Check if the "EmployeeData" sheet exists and read employee data from it
            if "EmployeeData" in workbook.sheetnames:
                employee_sheet = workbook["EmployeeData"]
                for row in employee_sheet.iter_rows(min_row=2, values_only=True):
                    emp_number = row[0]  # A2, A3, etc.
                    emp_name = row[1]    # B2, B3, etc.
                    emp_salary = float(row[8]) + float(row[9]) + float(row[10])  # I2 + J2 + K2, etc.

                    # Insert empty value for the type (this is the dropdown box)
                    row_data = [emp_number, emp_name, "", emp_salary, "", "", "", "", "", "", "", "", "", ""]
                    sheet.append(row_data)
            else:
                messagebox.showerror("Data Error", "EmployeeData sheet not found.")

            workbook.save(file_path)
            messagebox.showinfo("Sheet Created", f"The sheet '{sheet_name}' was created with the necessary headers and employee data.")
        else:
            sheet = workbook[sheet_name]

        # Display the sheet content
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            salary = float(row[3]) if row[3] else 0.0
            additional = float(row[6]) if row[6] else 0.0
            bonus = float(row[4]) if row[4] else 0.0
            deduction = float(row[7]) if row[7] else 0.0
            days_absent = float(row[8]) if row[8] else 0.0

            absent_deduction = (salary / 30) * days_absent
            total = (salary + additional + bonus) - (deduction + absent_deduction)

            # Map j to the actual column number
            for j, (value, width) in enumerate(zip(reversed(row), column_widths[::-1]), start=1):
                # Ensure value is not None and properly formatted
                value = "" if value is None else str(value)

                # Map column_index to the actual column number
                column_index = len(row) - j

                # Read-only fields
                if column_index in {0, 1, 3, 9, 10, 11}:
                    if column_index == 9:  # Column for "Absent Deduction"
                        value = f"{absent_deduction:.2f}"
                    elif column_index == 11:  # Column for "Total"
                        value = f"{total:.2f}"
                    salary_info_label = ctk.CTkEntry(self.salary_frame, font=("Arial", 16), justify="center", width=width)
                    salary_info_label.insert(0, value)
                    salary_info_label.configure(state="readonly")
                else:
                    if column_index == 2:  # 3rd column (dropdown box for "النوع")
                        options = ["ش", "م", "نقد"]
                        salary_info_label = ctk.CTkComboBox(self.salary_frame, values=options, font=("Arial", 16), width=width)
                        salary_info_label.set(value)
                    else:
                        salary_info_label = ctk.CTkEntry(self.salary_frame, font=("Arial", 16), justify="center", width=width)
                        salary_info_label.insert(0, value)
                        salary_info_label.configure(state="normal")

                salary_info_label.grid(row=i, column=j, padx=5, pady=5, sticky="e")

            # Add the checkbox on the far right, after the last column of data
            checkbox_var = ctk.StringVar(value="on")  # Default to "on"
            checkbox = ctk.CTkCheckBox(self.salary_frame, variable=checkbox_var, text="", width=20)
            checkbox.grid(row=i, column=len(row) + 1, padx=5, pady=5, sticky="e")

            # Store the checkbox state in a dictionary associated with the row number
            self.row_checkbox_dict[i] = checkbox_var

        # Adjust column configuration for better layout
        for j in range(len(headers) + 1):
            self.salary_frame.grid_columnconfigure(j, weight=1)
            self.salary_frame.grid_columnconfigure(len(headers) + 1, weight=0)  # Ensure no extra space

    def save_changes(self):
        """Save the changes made in the salary sheet back to the Excel file, based on checkbox state."""
        file_path = os.path.normpath(self.controller.excel_file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[self.selected_sheet_name]

        # Iterate over the rows in the grid (starting from row 0)
        for i in range(sheet.max_row - 1):  # We use sheet.max_row - 1 because grid rows start from 0
            checkbox_var = self.row_checkbox_dict.get(i + 2)  # Adjust to get the correct row
            if checkbox_var and checkbox_var.get() == "on":
                # Iterate over the columns for each row
                for j in range(1, sheet.max_column + 1):
                    widget = self.salary_frame.grid_slaves(row=i + 2, column=sheet.max_column + 1 - j)
                    if widget:
                        widget = widget[0]
                        if isinstance(widget, ctk.CTkEntry) or isinstance(widget, ctk.CTkComboBox):
                            new_value = widget.get()
                            sheet.cell(row=i + 2, column=j).value = new_value
                            print(f"Updated row {i + 2}, column {j} with value: {new_value}")

        workbook.save(file_path)
        messagebox.showinfo("Save", "Changes have been saved successfully.")
